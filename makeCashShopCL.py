import pandas as pd
import time
import os
import gc
import re
import numpy as np
import openpyxl as xl
from openpyxl.styles import Font, Alignment


#fileName = "유료상점.csv"
#target = pd.read_csv(fileName)
fileName = "유료상점.xlsx"
target = pd.read_excel(fileName,sheet_name = '유료상점',engine='openpyxl')
xlFileName = f"result_{time.strftime('%y%m%d_%H%M%S')}.xlsx"
#target["CashShop ID"] = target["CashShop ID"].replace(n,0)

target = target.replace('-',np.nan)
cashShopIdList = target.drop_duplicates(subset='CashShopID')["CashShopID"]
cashShopIdList = cashShopIdList.dropna(axis=0)
cashShopIdIndexList = cashShopIdList.index

totalCount = len(cashShopIdIndexList)
#print(cashShopIdList.astype(int))
print(f'추가 상품 개수 : {totalCount}')



class Sales():
    pkgID = ""
    pkgName = ""
    category = ""
    desc = ""
    info = ""
    price = ""
    bonus = ""
    itemList0 = ""
    itemList1 = ""
    itemList2 = None
    limit = ""

    def print(self):
        #print(self.itemName0)
        print(self.itemName1)
        #print(self.itemName2)

salesList = [Sales]

for j in range(0,totalCount):
    #print(cashShopIdIndexList[j], j+1)

    if (j+1) >= len(cashShopIdIndexList) :
        tempDf = target[cashShopIdIndexList[j]:]
    else :
        tempDf = target[cashShopIdIndexList[j]:cashShopIdIndexList[j+1]]
    tempDf = tempDf.reset_index()
    #for i in range(0,len(cashShopIdIndexList)):
    for i in range(0,1):
        a = Sales()
        a.pkgID = int(tempDf.loc[0,"CashShopID"])
        a.pkgName = tempDf.loc[0,"PkgName"] #+ "[귀속]"
        a.category = tempDf.loc[0,"Category"]
        a.price = str(tempDf.loc[0,"Price"])
        a.bonus = int(tempDf.loc[0,"Bonus"])
        a.limit = tempDf.loc[0,"Limit"]
        a.itemList0 = (tempDf.drop_duplicates(subset='Name0')['Name0'].dropna(axis=0) + "[귀속] " + tempDf['Count0'].dropna(axis=0) + "개").splitlines()
        #a.itemList1 = tempDf.drop_duplicates(subset='Name1')["Name1"].dropna(axis=0) + "[귀속] " + tempDf['Count1'].dropna(axis=0) + "개 (" + tempDf['ItemID1'].dropna(axis=0) +")"
        a.itemList1 = tempDf.drop_duplicates(subset='Name1')["Name1"].dropna(axis=0) + "[귀속] " + tempDf['Count1'].dropna(axis=0) + "개"
   

        b = tempDf.drop_duplicates(subset='Name1')[["Name1","Name2"]].dropna(axis=0).index
        #b[0]=4, b[1]=6
        #print(a.itemName1)
        # for i in range(0, len(b)):
            

        #     a.itemName1[x] += tempDf.loc[0,"Name2"] + "[귀속] " + str(tempDf.loc[0,"Count2"]) + "개"

        if salesList != None :
            salesList.append(a)
        else :
            salesList = a


        #print(salesList[j].pkgName)


    del tempDf
    gc.collect()


totalResult = pd.DataFrame()
print(len(salesList))

curRow = 0
for y in salesList:
    #print(y.pkgName)
    #print(i)
    #ta += 1
    #print(ta)
    #y=salesList[0]
    result = pd.DataFrame()
    #for i in range(0,100):  
    i = curRow
    result.loc[i,"Category2"] = y.pkgName + "\n" + str(y.pkgID)
    result.loc[i,"Category3"] = "이름"
    result.loc[i,"Check List"] = y.pkgName

#■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
    i += 1
    result.loc[i,"Category3"] = "카테고리"
    result.loc[i,"Check List"] = y.category
#■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
    i += 1
    result.loc[i,"Category3"] = "상세정보"
    desc0 = "\n".join(y.itemList0)
    desc0 = desc0.replace("다이아몬드[귀속]","다이아몬드")
    result.loc[i,"Check List"] = desc0

    i += 1
    desc1 = "\n".join(map(str, y.itemList1))
    desc1 = desc1.replace("nan\n","")
    desc1 = desc1.replace("\n","\n- ")
    result.loc[i,"Check List"] = "사용 시 다음 아이템 획득\n\n- "+desc1

    i += 1
    desc0 = desc0.replace("\n","\n- ")
    result.loc[i,"Check List"] = "<"+y.pkgName+"> 구성품 상세 정보\n- " + desc0

#■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■

    i += 1
    result.loc[i,"Category3"] = "상품슬롯"

    for item1 in y.itemList1 :
        if str(item1) != "nan" :
            result.loc[i,"Check List"] = str(item1) + " 관련 이미지 노출"
            i+=1

    result.loc[i,"Check List"] = "마일리지 : " + str(y.bonus)
    i+=1
    result.loc[i,"Check List"] = "구매 제한 : " + y.limit
    i+=1
    result.loc[i,"Check List"] = "구매 가격 : " + y.price
    
#■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■

    i += 1
    result.loc[i,"Category3"] = "아이템 구매"
    result.loc[i,"Check List"] = y.price + " 차감"
#■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■

    i += 1
    result.loc[i,"Category3"] = "마일리지"
    result.loc[i,"Check List"] =  str(y.bonus)+ " 마일리지 적립"
#■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
    i += 1
    result.loc[i,"Category3"] = "아이템 획득"
    result.loc[i,"Check List"] = y.pkgName + "상자[귀속] 인벤토리 획득"
#■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
    i += 1
    result.loc[i,"Category3"] = "아이템 사용"

    for item1 in y.itemList1 :
        if str(item1) != "nan" :
            result.loc[i,"Check List"] = "- " + str(item1) + " 획득 및 사용"
            i+=1
    
    i-=1
#■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
    i += 1
    result.loc[i,"Category3"] = "구매 제한"
    result.loc[i,"Check List"] = y.limit + " 구매 시 상품 슬롯 비활성화"
    i += 1
    result.loc[i,"Check List"] = "상품 슬롯 하단에 [구매 완료] 라벨 노출"
#■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
    curRow = i
    

    totalResult = pd.concat([totalResult,result], ignore_index=True)
    print(len(totalResult))

totalResult = totalResult.replace("NaN","")
totalResult = totalResult.replace("nan","")
totalResult = totalResult.replace(np.nan,"")

totalResult.to_excel(xlFileName, # directory and file name to write

            sheet_name = 'Sheet1', 

            na_rep = 'NaN', 

            float_format = "%.2f", 

            header = True, 

            #columns = ["group", "value_1", "value_2"], # if header is False

            index = True, 

            #index_label = "id", 

            startrow = 1, 

            startcol = 1, 

            #engine = 'xlsxwriter', 

            #freeze_panes = (2, 0)

            ) 



#a.category = "r"

#salesList[0].print()

# if __name__ == "__main__" : 
#     check_redraw_gacha(14)




#wb = xl.Workbook()
wb = xl.load_workbook(xlFileName,data_only = True)
sheetList = wb.sheetnames
#ws = xl.load_workbook()
ws = wb[sheetList[0]]
ws.column_dimensions['c'].width = 24
ws.column_dimensions['d'].width = 17
ws.column_dimensions['e'].width = 60










firstRow = 2
lastRow = ws.max_row
startRow_C = 0
startRow_D = 0
#endRow = 0

for i in range(firstRow, lastRow):
    #print(i)
    if ws['c'+str(i)].value is not None:
        if startRow_C == 0 :
            startRow_C = i
            #print(startRow)
        else :
            firstTargetCell =  "C"+str(startRow_C)
            mergeTargetCell = "C"+str(startRow_C)+":C"+str(i-1)
            ws.merge_cells(mergeTargetCell)
            # ws[firstTargetCell].alignment = Alignment(
            #     horizontal='center'
            #     ,vertical='top'
            #     ,wrap_text=True)
            startRow_C = i

    if ws['d'+str(i)].value is not None:
        if startRow_D == 0 :
            startRow_D = i
            #print(startRow)
        else :
            firstTargetCell =  "D"+str(startRow_D)
            mergeTargetCell = "D"+str(startRow_D)+":D"+str(i-1)
            ws.merge_cells(mergeTargetCell)
            # ws[firstTargetCell].alignment = Alignment(
            #     horizontal='center'
            #     ,vertical='top'
            #     ,wrap_text=True)
            startRow_D = i


    ws['c'+str(i)].alignment = Alignment(
        horizontal='center'
        ,vertical='top'
        ,wrap_text=True)
        
    ws['c'+str(i)].font = Font(size = 9, bold = True)
    ws['d'+str(i)].alignment = Alignment(
        horizontal='center'
        ,vertical='top'
        ,wrap_text=True)
    ws['d'+str(i)].font = Font(size = 9, bold = True)
    ws['e'+str(i)].alignment = Alignment(
        horizontal='left'
        ,vertical='top'
        ,wrap_text=True)
    ws['e'+str(i)].font = Font(size = 9, bold = False)



#예외 마지막 셀병합
ws.merge_cells("C"+str(startRow_C)+":C"+str(lastRow))
ws.merge_cells("C"+str(startRow_D)+":C"+str(lastRow))




# ws.column_dimensions['c'].alignment = Alignment(
#     horizontal='center'
#     ,vertical='top'
#     ,wrap_text=True)

# ws.column_dimensions['d'].alignment = Alignment(
#     horizontal='center'
#     ,vertical='top'
#     ,wrap_text=True)

# ws.column_dimensions['e'].alignment = Alignment(
#     horizontal='left'
#     ,vertical='top'
#     ,wrap_text=True)

wb.save(xlFileName)
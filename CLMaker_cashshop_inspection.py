import pandas as pd
import time
import os
import gc
import re
import numpy as np
import openpyxl as xl
from openpyxl.styles import Font, Alignment, Color
from tqdm import tqdm
from openpyxl.styles.colors import BLACK
import datetime
from openpyxl.styles import Font, Color
#from openpyxl.styles.colors import RED
from openpyxl.utils import get_column_letter
#cashShopIdIndexList = cashShopIdList.index

cashShopDir = "./CL_CashShop_Inspection"
if not os.path.isdir(cashShopDir) :
    os.mkdir(cashShopDir)
tempDir = "./temp"
if not os.path.isdir(tempDir) :
    os.mkdir(tempDir)

xlFileName = f"./CL_CashShop_Inspection/result_{time.strftime('%y%m%d_%H%M%S')}.xlsx"
tempCsvName = f"./temp/tempCsv.csv"

def dateCheck(start_date, end_date):
    today = datetime.date.today()

    
    if start_date.date() == today :
        return "판매 시작"
    elif start_date.date() < today < end_date.date():
        return "판매 유지"
    elif today == end_date.date():
        return "판매 종료"
    else:
        return "판매 제외"

class Sales():

    
    def __init__(self) :

        self.pkgID = ""
        self.pkgName = ""
        self.category = ""
        self.desc = ""
        self.info = ""
        self.price = ""
        self.bonus = ""
        self.itemList0 = ""
        self.itemList1 = ""
        self.itemList2 =""
        self.limit = ""
        self.server = ""
        self.startDate = ""
        self.endDate = ""
        self.itemList0 = []
        self.itemList1 = []
        self.itemList2 =[]

        #별도 저장값
        self.salesCheck = ""

class Item():

    def __init__(self) :
        self.name = ""
        self.id = ""
        self.innerItemList = ""


def extract_data_cashshop(fileName):

#CSV 읽기
    #target = pd.read_csv(fileName)

#XLSX 읽기
    tempTarget = pd.read_excel(fileName,engine='openpyxl', na_values = "")
    tempTarget.to_csv(tempCsvName, encoding='cp949')
    target = pd.read_csv(tempCsvName, encoding='cp949')





    # fileName = "유료상점.xlsx"
    # target = pd.read_excel(fileName,sheet_name = '유료상점',engine='openpyxl')
    #target["CashShop ID"] = target["CashShop ID"].replace(n,0)

    target = target.replace('-',np.nan)
    cashShopIdList = target.drop_duplicates(subset='CashShopID')["CashShopID"]
    cashShopIdList = cashShopIdList.dropna(axis=0)
    cashShopIdIndexList = cashShopIdList.index

    totalCount = len(cashShopIdIndexList)
    #print(cashShopIdList.astype(int))
    #print(f'추가 상품 개수 : {totalCount}')

    gachaItemIndexList = target[["ItemID1","ItemID2"]].dropna(axis=0).index
    #cashShopIdList = cashShopIdList.dropna(axis=0)
    #print(gachaItemIndexList)
    salesList = [Sales] 
    #salesList : list[Sales]
    salesList.clear()
    print("데이터 추출 중...")
    tqdmCount2 =0
    for j in tqdm(range(0,totalCount)):
        tqdmCount2 +=1
        #print(cashShopIdIndexList[j], j+1)

        if (j+1) >= len(cashShopIdIndexList) :
            tempDf = target[cashShopIdIndexList[j]:]
        else :
            tempDf = target[cashShopIdIndexList[j]:cashShopIdIndexList[j+1]]
        tempDf = tempDf.reset_index()
        #for i in range(0,len(cashShopIdIndexList)):
        #for i in range(0,1):

        a = Sales()
        a.pkgID = int(tempDf.loc[0,"CashShopID"])
        a.pkgName = tempDf.loc[0,"PkgName"] #+ "[귀속]"
        a.category = tempDf.loc[0,"Category"]
        a.price = str(tempDf.loc[0,"Price"])
        a.bonus = int(tempDf.loc[0,"Bonus"])
        a.limit = tempDf.loc[0,"Limit"]

        for k in range(len(tempDf)):
            #print(len(tempDf))
            if not pd.isnull(tempDf.iloc[k]['Name0']):
                itemName = tempDf.iloc[k]['Name0']
                itemCount = tempDf.iloc[k]['Count0']
                a.itemList0.append(f"{itemName}[귀속] {int(itemCount)}개")
                #print(a.itemList0)

        for k in range(len(tempDf)):
            if not pd.isnull(tempDf.iloc[k]['Name1']):
                itemName = tempDf.iloc[k]['Name1']
                itemCount = tempDf.iloc[k]['Count1']
                try: 
                    a.itemList1.append(f"{itemName}[귀속] {int(itemCount)}개")
                except:
                    a.itemList1.append(f"{itemName}[귀속] {(itemCount)}개")

        #a.itemList0 = tempDf.drop_duplicates(subset='Name0')['Name0'].dropna(axis=0) + "[귀속] " + str(tempDf['Count0'].dropna(axis=0)) + "개"#.splitlines()
        #a.itemList0 = f"{str(tempDf.drop_duplicates(subset='Name0')['Name0'])}[귀속] {str(tempDf['Count0'])}개".splitlines()
        #a.itemList1 = tempDf.drop_duplicates(subset='Name1')["Name1"].dropna(axis=0) + "[귀속] " + tempDf['Count1'].dropna(axis=0) + "개"
        #a.itemList1 = tempDf.drop_duplicates(subset='Name1')["Name1"].dropna(axis=0) + "[귀속] " + str(tempDf['Count1'].dropna(axis=0)) + "개"
        a.server = tempDf.loc[0,"Server"]
        b = tempDf.drop_duplicates(subset='Name1')[["Name1","Name2"]].dropna(axis=0).index
        #b[0]=4, b[1]=6
        #print(a.itemName1)
        # for i in range(0, len(b)):

        a.startDate = pd.to_datetime(tempDf.loc[0,"StartDate"])

        if "상시" in str(tempDf.loc[0,"EndDate"]) :
            a.endDate = "2099-12-31 00:00:00"
        else: 
            a.endDate = pd.to_datetime(tempDf.loc[0,"EndDate"])
            
        #print(a.startDate,a.endDate)

        #     a.itemName1[x] += tempDf.loc[0,"Name2"] + "[귀속] " + str(tempDf.loc[0,"Count2"]) + "개"

        #if dateCheck(a.startDate,a.endDate)

        a.salesCheck = dateCheck(a.startDate,a.endDate)



        if salesList != None :
            salesList.append(a)
        else :
            salesList = a



        #print(a.itemList0)
        #print(a.itemList1)
        #print(salesList[j].pkgName)
        #a.itemList0.clear()
        #a.itemList1.clear()
        #a.itemList2.clear()
        del a,tempDf
        gc.collect()

        #print(a.itemList2)
        #del tempDf
        #gc.collect()




    salesList.sort(key =lambda a: a.salesCheck)
    #for s in salesList:
    #    print(s.server)
    return salesList

def write_data_cashshop(salesList : list[Sales]):
    totalResult = pd.DataFrame()
#print(len(salesList))

    curRow = 0
    count = 0
    tqdmCount0=0
    print("데이터 쓰는 중...")
    for y in tqdm(salesList):
        tqdmCount0+=1
        y : Sales
        count += 1
        result = pd.DataFrame()

        i = curRow
        result.loc[i,"Category1"] = "유료 상품"
        result.loc[i,"Category2"] = y.salesCheck
        result.loc[i,"Category3"] = y.pkgName
        #result.loc[i,"Check List"] = y.pkgName

    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
        #i += 1
        #result.loc[i,"Category3"] = "카테고리"
        #result.loc[i,"Check List"] = y.category
    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
        #i += 1
        #result.loc[i,"Category3"] = "상세정보"
    
        info_0 = f'{y.category} / {y.price} / {y.bonus} / {y.limit}'

        info_1 = "\n".join(y.itemList0)
        info_1 = info_1.replace("다이아몬드[귀속]","다이아몬드")
        #result.loc[i,"Check List"] = desc0

        #i += 1
        info_2 = "\n".join(map(str, y.itemList1))
        info_2 = info_2.replace("nan\n","")
        info_2 = info_2.replace("\n","\n- ")
        info_2 = "사용 시 다음 아이템 획득\n- "+info_2

        #i += 1
        #desc0 = desc0.replace("\n","\n- ")
        #result.loc[i,"Check List"] = "<"+y.pkgName+"> 구성품 상세 정보\n- " + desc0

        info_3 = f'* 상세정보 및 패키지 상자 구성품 내 [귀속] 노출 확인\n* 패키지 이미지 내 구성품 관련 이미지 노출 확인'

        result.loc[i,"Check List"] = f'{info_0}\n\n{info_1}\n\n{info_2}\n\n{info_3}'

    # #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■

    #     i += 1
    #     result.loc[i,"Category3"] = "상품슬롯"

    #     for item1 in y.itemList1 :
    #         if str(item1) != "nan" :
    #             result.loc[i,"Check List"] = str(item1) + " 관련 이미지 노출"
    #             i+=1

    #     if int(y.bonus) == 0 :
    #         result.loc[i,"Check List"] =  "마일리지 미노출"
    #     else :            
    #         result.loc[i,"Check List"] =  "마일리지 : " + str(y.bonus)+ " 적립"
    #     i+=1
    #     result.loc[i,"Check List"] = "구매 제한 : " + y.limit
    #     i+=1
    #     result.loc[i,"Check List"] = "구매 가격 : " + y.price
        
    # #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■

    #     i += 1
    #     result.loc[i,"Category3"] = "아이템 구매"

    #     if "원" in y.price :
    #         result.loc[i,"Check List"] = f"결제 모듈 내 {y.pkgName} 노출"
    #         i += 1
    #         result.loc[i,"Check List"] = f"결제 모듈 내 {y.price} 노출"
    #         i += 1
    #         result.loc[i,"Check List"] = f"결제 완료 시 보관함으로 획득"
    #     else :
    #         result.loc[i,"Check List"] = y.price + " 차감"
    # #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■

    #     i += 1
    #     result.loc[i,"Category3"] = "마일리지"

    #     if int(y.bonus) == 0 :
    #         result.loc[i,"Check List"] =  "미노출"
    #     else :            
    #         result.loc[i,"Check List"] =  "마일리지 : " + str(y.bonus)+ " 적립"

    # #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
    #     i += 1
    #     result.loc[i,"Category3"] = "아이템 획득"
    #     result.loc[i,"Check List"] = y.pkgName + "상자[귀속] 인벤토리 획득"
    # #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
    #     i += 1
    #     result.loc[i,"Category3"] = "아이템 사용"

    #     for item1 in y.itemList1 :
    #         if str(item1) != "nan" :
    #             result.loc[i,"Check List"] = "- " + str(item1) + " 획득 및 사용"
    #             i+=1
        
    #     i-=1
    # #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
    #     i += 1
    #     result.loc[i,"Category3"] = "구매 제한"
    #     result.loc[i,"Check List"] = y.limit + " 구매 시 상품 슬롯 비활성화"
    #     i += 1
    #     result.loc[i,"Check List"] = "상품 슬롯 하단에 [구매 완료] 라벨 노출"
    # #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
    #     curRow = i
        
        if y.salesCheck != "판매 제외" :
            totalResult = pd.concat([totalResult,result], ignore_index=True)
        #print(len(totalResult))

    totalResult = totalResult.replace("NaN","")
    totalResult = totalResult.replace("nan","")
    totalResult = totalResult.replace(np.nan,"")

    totalResult.to_excel(xlFileName, # directory and file name to write

                sheet_name = 'Sheet1', 

                na_rep = 'NaN', 

                float_format = "%.2f", 

                header = True, 

                #columns = ["group", "value_1", "value_2"], # if header is False

                index = True, 

                #index_label = "id", 

                startrow = 0, 

                startcol = 0, 

                #engine = 'xlsxwriter', 

                #freeze_panes = (2, 0)

                ) 


def postprocess_cashshop():
    wb = xl.load_workbook(xlFileName,data_only = True)
    sheetList = wb.sheetnames
    ws = wb[sheetList[0]]
    ws.column_dimensions['b'].width = 17
    ws.column_dimensions['c'].width = 24
    ws.column_dimensions['d'].width = 17
    ws.column_dimensions['e'].width = 60

    firstRow = 2
    lastRow = ws.max_row
    startRow_B =0
    startValue_B =""
    startRow_C = 0
    startRow_D = 0

    tqdmCount1 = 0
    print("엑셀 서식 처리중...")
    for i in tqdm(range(firstRow, lastRow+1)):
        tqdmCount1+=1
        #print(i)
        if (ws['b'+str(i)].value is not None) :
            if startRow_B == 0  :
                startRow_B = i
                startValue_B = ws['b'+str(i)].value
                #print(startRow_B)
            else :
                #firstTargetCell =  "C"+str(startRow_C)
                if ( ws['b'+str(i)].value != startValue_B) :
                    mergeTargetCell = "B"+str(startRow_B)+":B"+str(i-1)
                    ws.merge_cells(mergeTargetCell)
                    startValue_B = ws['b'+str(i)].value
                    startRow_B = i

        if ws['c'+str(i)].value is not None:
            if startRow_C == 0 :
                startRow_C = i
                #print(startRow)
            else :
                firstTargetCell =  "C"+str(startRow_C)
                mergeTargetCell = "C"+str(startRow_C)+":C"+str(i-1)
                ws.merge_cells(mergeTargetCell)
                startRow_C = i

        if ws['d'+str(i)].value is not None:
            if startRow_D == 0 :
                startRow_D = i
                #print(startRow)
            else :
                firstTargetCell =  "D"+str(startRow_D)
                mergeTargetCell = "D"+str(startRow_D)+":D"+str(i-1)
                ws.merge_cells(mergeTargetCell)
                startRow_D = i


        ws['b'+str(i)].alignment = Alignment(
            horizontal='center'
            ,vertical='top'
            ,wrap_text=True)
        ws['b'+str(i)].font = Font(size = 9, bold = True)
        ws['c'+str(i)].alignment = Alignment(
            horizontal='center'
            ,vertical='top'
            ,wrap_text=True)
        ws['c'+str(i)].font = Font(size = 9, bold = True)
        ws['d'+str(i)].alignment = Alignment(
            horizontal='center'
            ,vertical='top'
            ,wrap_text=True)
        ws['d'+str(i)].font = Font(size = 9, bold = True)
        ws['e'+str(i)].alignment = Alignment(
            horizontal='left'
            ,vertical='top'
            ,wrap_text=True)
        ws['e'+str(i)].font = Font(size = 9, bold = False)
        
        #ws['e'+str(i)].value = process_temp_str(str(ws['e'+str(i)].value))


    #예외 마지막 셀병합
    ws.merge_cells("B"+str(startRow_B)+":B"+str(lastRow))
    ws.merge_cells("C"+str(startRow_C)+":C"+str(lastRow))
    ws.merge_cells("D"+str(startRow_D)+":D"+str(lastRow))

    ws = highlight_belonging(ws)

    wb.save(xlFileName)

def process_temp_str(temp_str):
    # Define the font style for the red asterisk
    red_asterisk_font = Font(color=Color('FF0000'), bold=True)
    # Define the font style for the rest of the cell
    normal_font = Font(color=Color('000000'))

    # Split the temp_str value by newline characters
    temp_str_lines = temp_str.split("\n")

    # Process each line of the temp_str value
    result_lines = []
    for line in temp_str_lines:
        # Check if the line starts with an asterisk
        if line.startswith("*"):
            # Apply the red font to the asterisk and the rest of the line
            result_lines.append((line[:1] + line[1:].strip(), red_asterisk_font))
        else:
            # Apply the normal font to the entire line
            result_lines.append((line, normal_font))

    # Join the processed lines and return the result
    #return "\n".join(result_lines)
    #print(result_lines)
    return "\n".join([line[0] for line in result_lines])


def highlight_belonging(ws):
    # Define the font style for the highlighted text
    highlight_font = Font(color=Color('FF0000'), bold=True)

    # Iterate over all cells in the worksheet
    for row in ws.iter_rows():
        for cell in row:
            # Check if the cell contains the word "belonging"
            if "귀속" in str(cell.value).lower():
                # Split the cell value by the word "belonging"
                parts = str(cell.value).lower().split("귀속")
                # Concatenate the parts with the highlighted "belonging" in between
                highlighted_value = "귀속".join([part.strip() for part in parts])
                # Apply the highlight font to the "belonging" text
                cell.font = highlight_font
                # Set the cell value to the highlighted value
                cell.value = highlighted_value

    # Auto-fit the columns to adjust to the new cell values
    for column in ws.columns:
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].auto_size = True

    # Return the worksheet object
    return ws


if __name__ == "__main__":

    #print("┃  R2M CASH SHOP CL MAKER  ┃")
    #fileName = input("> 데이터파일명 입력(엔터:유료상점DATA.csv) : ")
    fileName = ""
    if fileName == "":
        #fileName = "유료상점DATA.csv"
        fileName = "유료상점DATA_0.xlsx"

    while not os.path.isfile(fileName) :
        fileName = input("> Insert csv file name : ")


    salesList = extract_data_cashshop(fileName)
    write_data_cashshop(salesList)
    postprocess_cashshop()

    print("생성완료")
    input("종료하려면 엔터키 입력...")
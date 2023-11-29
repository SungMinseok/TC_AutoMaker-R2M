import pandas as pd
import time
import os
import gc
import re
import numpy as np
import openpyxl as xl
from openpyxl.styles import Font, Alignment, Color
from tqdm import tqdm
from openpyxl.styles.colors import BLACK
import datetime
from openpyxl.styles import Font, Color
#from openpyxl.styles.colors import RED
from openpyxl.utils import get_column_letter
#cashShopIdIndexList = cashShopIdList.index
clType = ""
cashShopDir = ""
tempDir = "./temp"
if not os.path.isdir(tempDir) :
    os.mkdir(tempDir)
tempCsvName = f"./temp/tempCsv.csv"

#xlFileName = ""
#tcStartDate = ""

idList = [int]

def dateCheck(start_date, end_date, today = datetime.date.today()):
    
    if start_date.date() == today :
        return "판매 시작"
    elif start_date.date() < today < end_date.date():
        return "판매 유지"
    elif today == end_date.date():
        return "판매 종료"
    # elif start_date.date() >= today - datetime.timedelta(days=1)  :
    #     return "판매 전"
    elif start_date.date() >= today  :
        return "판매 전"
    else:
        return "판매 제외"

class Sales():
    def __init__(self) :

        self.pkgID = ""
        self.pkgName = ""
        self.category = ""
        self.desc = ""
        self.info = ""
        self.price = ""
        self.bonus = ""
        self.itemList0 = ""
        self.itemList1 = ""
        self.itemList2 =""
        self.limit = ""
        self.server = ""
        self.startDate = ""
        self.endDate = ""
        self.itemList0 = []
        self.itemList1 = []
        self.itemList2 =[]
        self.order = -1

        #별도 저장값
        self.salesCheck = ""

class Event():
    def __init__(self) :

        self.id = ""
        self.type = ""
        self.name = ""
        self.quest_desc = ""
        self.item_list_0 = []
        self.item_list_1 = []
        self.item_list_2 = []
        # self.item_id_0 = ""
        # self.item_name_0 = ""
        # self.item_count_0 = ""
        # self.item_id_1 = ""
        # self.item_name_1 = ""
        # self.item_count_1 = ""
        self.craft_price = ""
        self.craft_ingred = ""
        self.limit = ""
        self.server = ""
        self.start_date = ""
        self.end_date = ""

        #별도 저장값
        self.open_check = ""

class Item():

    def __init__(self) :
        self.name = ""
        self.id = ""
        self.innerItemList = ""
        self.rate = ""







def extract_data_cashshop(fileName, tcStartDate):

#CSV 읽기
    #target = pd.read_csv(fileName)

#XLSX 읽기
    # tempTarget = pd.read_excel(fileName,engine='openpyxl', na_values = "")
    # tempTarget.to_csv(tempCsvName, encoding='cp949')
    # target = pd.read_csv(tempCsvName, encoding='cp949')
    #target = pd.read_excel(fileName,engine='openpyxl', na_values = "")


    # # 모든 시트 이름 가져오기
    # sheet_names = pd.read_excel(fileName, sheet_name=None).keys()
    # sheet_names1 = pd.read_excel(fileName, sheet_name=None)

    # # 모든 시트의 데이터프레임을 리스트에 저장
    # dfs = []
    # headerCheck = False
    # for sheet_name in sheet_names:
    #     if not headerCheck :
    #         headerCheck = True
    #         df = pd.read_excel(fileName, sheet_name=sheet_name, header = 0, na_values="")
    #     else :
    #         df = pd.read_excel(fileName, sheet_name=sheet_name,header = None,  na_values="")
    #         #df = df[1:]
    #     dfs.append(df)

    # # 모든 데이터프레임을 연속해서 연결하여 하나의 데이터프레임으로 만듦
    # target = pd.concat(dfs, axis = 0, ignore_index=True)
    target = pd.DataFrame()


    #sheet_names = pd.read_excel(fileName, sheet_name=None).keys()
    sheet_names = pd.ExcelFile(fileName).sheet_names

    #시트 너무 많으면 오래걸려서 시트 개수 제한 n개
    try:
        sheet_names = sheet_names[:3]
    except:
        pass
    for i, sheet_name in enumerate(sheet_names):
        
        curDf = pd.read_excel(fileName, sheet_name=sheet_name, na_values=np.nan)

        #target = target.append(curDf, ignore_index = True)
        target = pd.concat([target, curDf], ignore_index=True)
        #try:
        #target = target.replace('　', '', regex=True, inplace=True)
        #target = target.replace('', pd.NA, inplace=True)
            #target = target.replace('-', pd.NA, inplace=True)
        

        del curDf
        gc.collect()

    
    # 모든 데이터프레임을 연속해서 연결하여 하나의 데이터프레임으로 만듦
    #target = pd.concat(dfs, ignore_index=True)



    # fileName = "유료상점.xlsx"
    # target = pd.read_excel(fileName,sheet_name = '유료상점',engine='openpyxl')
    #target["CashShop ID"] = target["CashShop ID"].replace(n,0)

    #target = target.replace('-',np.nan)
    #cashShopIdList = target.drop_duplicates(subset='CashShopID')["CashShopID"]
    cashShopIdList = target["CashShopID"].dropna(axis=0)
    cashShopIdIndexList = cashShopIdList.index

    totalCount = len(cashShopIdIndexList)
    #print(cashShopIdList.astype(int))
    #print(f'추가 상품 개수 : {totalCount}')

    #gachaItemIndexList = target[["ItemID1","ItemID2"]].dropna(axis=0).index
    #cashShopIdList = cashShopIdList.dropna(axis=0)
    #print(gachaItemIndexList)
    salesList = [Sales] 
    #salesList : list[Sales]
    salesList.clear()
    print("데이터 추출 중...")
    for j in tqdm(range(0,totalCount)):
        #print(cashShopIdIndexList[j], j+1)

        if (j+1) >= len(cashShopIdIndexList) :
            tempDf = target[cashShopIdIndexList[j]:]
        else :
            tempDf = target[cashShopIdIndexList[j]:cashShopIdIndexList[j+1]]
        tempDf = tempDf.reset_index()
        #for i in range(0,len(cashShopIdIndexList)):
        #for i in range(0,1):
        tempDf = tempDf.applymap(lambda x: x.strip() if isinstance(x, str) else x)
        tempDf = tempDf.replace('-',np.nan)

        a = Sales()
        a.pkgID = int(tempDf.loc[0,"CashShopID"])
        a.pkgName = tempDf.loc[0,"PkgName"] #+ "[귀속]"
        a.category = str(tempDf.loc[0,"Category"])
        a.order = tempDf.loc[0,"Order"]
        a.price = str(tempDf.loc[0,"Price"])
        try:
            a.bonus = int(tempDf.loc[0,"Bonus"])
        except:
            a.bonus = 0
        a.limit = tempDf.loc[0,"Limit"]

        if a.pkgID == 1100053 :
            print('adasdsa')

        for k in range(len(tempDf)):
            #print(len(tempDf))
            if not pd.isnull(tempDf.iloc[k]['Name0']):
                if tempDf.iloc[k]['Name0'] == "" :
                    continue
                itemName = tempDf.iloc[k]['Name0']
                itemCount = tempDf.iloc[k]['Count0']
                try : 
                    a.itemList0.append(f"{itemName}[귀속] {int(itemCount)}개")
                except:
                    a.itemList0.append(f"{itemName}[귀속] {(itemCount)}개")

                #a.itemList0.sort()
                coin_items0 = [item for item in a.itemList0 if '로얄 코인' == item]
                non_coin_items0 = [item for item in a.itemList0 if '로얄 코인' != item]
                a.itemList0 = coin_items0 + non_coin_items0

        for k in range(len(tempDf)):
            if not pd.isnull(tempDf.iloc[k]['Name1']):
                if tempDf.iloc[k]['Name1'] == "-" :
                    continue
                itemName = tempDf.iloc[k]['Name1']
                itemCount = tempDf.iloc[k]['Count1']
                try: 
                    a.itemList1.append(f"{itemName}[귀속] {int(itemCount)}개")
                except:
                    a.itemList1.append(f"{itemName}[귀속] {(itemCount)}개")
                
                #a.itemList1.sort()
                coin_items1 = [item for item in a.itemList1 if '로얄 코인' == item]
                non_coin_items1 = [item for item in a.itemList1 if '로얄 코인' != item]
                a.itemList1 = coin_items1 +non_coin_items1
        
        for k in range(len(tempDf)):
            if not pd.isnull(tempDf.iloc[k]['Name2']):
                itemName = tempDf.iloc[k]['Name2']
                itemCount = tempDf.iloc[k]['Count2']
                try: 
                    a.itemList2.append(f"{itemName}[귀속] {int(itemCount)}개")
                except:
                    a.itemList2.append(f"{itemName}[귀속] {(itemCount)}개")

        # for k in range(len(tempDf)):
        #     if not pd.isnull(tempDf.iloc[k]['Name1']):
        #         for l in range(len(tempDf)):
                    
        #             if l == 0 :
        #                 a.itemList0[k] += '\n사용 시 다음 아이템 획득'
        #             a.itemList0[k] += f'\n{a.itemList1[l]}'

        #             if not pd.isnull(tempDf.iloc[k+1]['Name0']):
        #                 break
                    



        a.server = str(tempDf.loc[0,"Server"]).strip()
       
        a.startDate = pd.to_datetime(tempDf.loc[0,"StartDate"])

        if "상시" in str(tempDf.loc[0,"EndDate"]) :
            a.endDate = datetime.datetime.strptime("2099-12-31 00:00:00",'%Y-%m-%d %H:%M:%S')
        else: 
            try: 
                a.endDate = pd.to_datetime(tempDf.loc[0,"EndDate"])
            except : 
                print(f"상품 시작/종료 시간 입력 오류 : 'yyyy-mm-dd' 만 입력되어야 함. 시간이 써있지 않은지 확인 필요\n상품 : {a.pkgID} | {a.pkgName}" )
                os.system('pause')
                return
          
        #if fileType == "0" :#TC
            #startDate = datetime.datetime.strptime(tcStartDate, '%Y-%m-%d')
            #a.salesCheck = dateCheck(a.startDate,a.endDate,startDate.date())
        #elif fileType == "1" :#점검
        startDate = datetime.datetime.strptime(tcStartDate, '%Y-%m-%d')
        a.salesCheck = dateCheck(a.startDate,a.endDate,startDate.date())



        if salesList != None :
            salesList.append(a)
        else :
            salesList = a

        #print(a)

        del a,tempDf
        gc.collect()

    #salesList.sort(key =lambda a: (a.server,a.category))
    return salesList


def write_data_cashshop(salesList : list[Sales], resultPath = "유료상점_TestCase"):
    '''
    유료상점_테스트케이스_작성용
    '''
    totalResult = pd.DataFrame()
#print(len(salesList))

    #salesList.sort(key =lambda a: (a.server,a.category,a.pkgID))

    category_order = ['시즌 뽑기', '기간한정상품', '패키지', '카드', '재화', '이벤트', '마일리지']
    
    for sale in salesList:
        print(f'{sale.pkgName}|{sale.server}|{sale.salesCheck}|{sale.category}|{sale.order}')
    try:
        #    sale.salesCheck = float(sale.salesCheck)
        #salesList.sort(key=lambda a: (a.server, a.salesCheck, a.category, a.order))
        salesList.sort(key=lambda a: (
            #a.salesCheck,
            a.server,
            category_order.index(a.category) if a.category in category_order else float('inf'),
            a.order
        ))

    except Exception as e:
        print(e)
        print("정렬에 문제 발생... 표에 제대로 입력됐는지 확인 필요...")


    curRow = 0
    count = 0
    tqdmCount0=0
    print("데이터 쓰는 중...")
    for y in tqdm(salesList):
        tqdmCount0+=1
        y : Sales
        count += 1
        result = pd.DataFrame()

        i = curRow
        result.loc[i,"Category1"] = y.server
        result.loc[i,"Category2"] = y.pkgName + "\n" + str(y.pkgID)
        result.loc[i,"Category3"] = "이름"
        result.loc[i,"Check List"] = y.pkgName

    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
        i += 1
        result.loc[i,"Category3"] = "카테고리"
        result.loc[i,"Check List"] = y.category
    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■

        if y.pkgID == 1100043 :
            print("test")

        if y.category == "" :
            i += 1
            result.loc[i,"Category3"] = "상세정보 팝업"

        else :
            i += 1
            result.loc[i,"Category3"] = "상세정보 팝업"
            desc0 = ""
            if len(y.itemList0) != 0 :
                for item in y.itemList0 :
                    if '다이아몬드' in item :
                        item = item.replace("다이아몬드[귀속]","다이아몬드")
                        #item = item.replace("다이아[귀속]","다이아몬드")
                        item = f'{item}\nR2M에서 사용되는 유료 재화 입니다.\n\n'
                    
                    desc0 = "\n".join(map(str, y.itemList0))
                    
                    # if y.category == '시즌 뽑기' :
                    #     desc0 = "\n".join(map(str, y.itemList0))
                    # else :
                    #     desc0 += item
                if len(y.itemList0) != 0 and len(y.itemList1) != 0 :
                    desc0 += f'\n사용 시 다음 아이템 획득'
                elif len(y.itemList1) != 0 and len(y.itemList2) != 0 :
                    desc0 += f'\n사용 시 다음 아이템 중 1종 획득'
                elif len(y.itemList1) == 0 and len(y.itemList2) != 0 :
                    desc0 += f'\n사용 시 다음 아이템 중 1종 획득'
                result.loc[i,"Check List"] = desc0


                if len(y.itemList1) != 0 :
                    i += 1
                    #desc1 = f'사용 시 다음 아이템 획득'
                    desc1 = "\n".join(map(str, y.itemList1))
                    desc1 = desc1.replace("nan\n","")
                    desc1 = desc1.replace("코인[귀속]","코인")
                    result.loc[i,"Check List"] = desc1

                if len(y.itemList2) != 0 :
                    i += 1
                    #desc1 = f'사용 시 다음 아이템 획득'
                    desc1 = "\n".join(map(str, y.itemList2))
                    desc1 = desc1.replace("nan\n","")
                    desc1 = desc1.replace("코인[귀속]","코인")
                    result.loc[i,"Check List"] = desc1






            else :                 
                if len(y.itemList1) != 0 :
                    desc0 = "\n".join(map(str, y.itemList1))
                    if len(y.itemList2) != 0 :
                        desc0 += f'\n사용 시 다음 아이템 중 1종 획득'
                    result.loc[i,"Check List"] = desc0

                else:
                    if '뽑기' not in str(y.pkgName) :
                        result.loc[i,"Check List"] = f'{y.pkgName} 상자[귀속]'
                    else :
                        result.loc[i,"Check List"] = f'{y.pkgName} 관련 뽑기 내역 노출'
            

                if len(y.itemList2) != 0 :
                    i += 1
                    #desc2 = f'사용 시 다음 아이템 중 1종 획득'
                    desc2 = "\n".join(map(str, y.itemList2))
                    desc2 = desc2.replace("nan\n","")
                    desc2 = desc2.replace("코인[귀속]","코인")
                    result.loc[i,"Check List"] = desc2


    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■

        i += 1
        result.loc[i,"Category3"] = "상품슬롯 이미지"

        if y.pkgID == 80060 :
            print('didiid')

        if len(y.itemList0) != 0 :
            for item0 in y.itemList0 :
                if '다이아몬드' in str(item0):
                    result.loc[i,"Check List"] = "다이아몬드 이미지 노출"
                    i+=1
                else :
                    result.loc[i,"Check List"] = f"{item0} 관련 이미지 노출"
                    i+=1
        elif len(y.itemList1) != 0 :
            for item0 in y.itemList0 :
                #i += 1
                if '다이아몬드' in str(item0):
                    result.loc[i,"Check List"] = "다이아몬드 이미지 노출"
                    i+=1

            for item1 in y.itemList1 :
                #i += 1
                if str(item1) != "nan" :
                    item1 = item1.replace("코인[귀속]","코인")
                    result.loc[i,"Check List"] = str(item1) + " 이미지 노출"
                    i+=1
        else:
            result.loc[i,"Check List"] = f'{y.pkgName} 관련 이미지 노출'
        i += 1
        result.loc[i,"Category3"] = "상품슬롯 정보"
        if int(y.bonus) == 0 :
            result.loc[i,"Check List"] =  "마일리지 미노출"
        else :            
            result.loc[i,"Check List"] =  "마일리지 : " + str(y.bonus)+ " 적립"
        i+=1
        if str(y.limit) == 'nan' or str(y.limit) == '무제한':
            result.loc[i,"Check List"] = "구매 제한 없음"
        else :            
            result.loc[i,"Check List"] = "구매 제한 : " + str(y.limit)
        i+=1
        result.loc[i,"Check List"] = "구매 가격 : " + y.price
        
    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■

        i += 1
        result.loc[i,"Category3"] = "아이템 구매"

        if "원" in y.price or "TWD" in y.price:
            result.loc[i,"Check List"] = f"결제 모듈 내 {y.pkgName} 노출"
            i += 1
            result.loc[i,"Check List"] = f"결제 모듈 내 {y.price} 노출"
            i += 1
            result.loc[i,"Check List"] = f"결제 완료 시 보관함으로 획득"
        else :
            result.loc[i,"Check List"] = y.price + " 차감"
    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■

        i += 1
        result.loc[i,"Category3"] = "마일리지"

        if int(y.bonus) == 0 :
            result.loc[i,"Check List"] =  "미획득"
            i -= 1
        else :            
            result.loc[i,"Check List"] =  "마일리지 : " + str(y.bonus)+ " 적립"

    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
        i += 1
        if "원" in y.price or "TWD" in y.price:
            result.loc[i,"Category3"] = "보관함 수령"
        elif y.category == "시즌 뽑기" or '뽑기' in y.pkgName:
            result.loc[i,"Category3"] = "뽑기 획득"
        else:
            result.loc[i,"Category3"] = "아이템 획득"
        if len(y.itemList1) != 0 :
                
            for item0 in y.itemList0 :
                if '다이아' not in item0 :
                    result.loc[i,"Check List"] = f'{item0} 인벤토리로 획득'#y.pkgName + " 인벤토리로 획득"
                else:
                    item0 = item0.replace('[귀속]','')
                    result.loc[i,"Check List"] = f'{item0} 지급'#y.pkgName + " 인벤토리로 획득"

                i += 1
        else:
            
            if y.category == "시즌 뽑기":
                result.loc[i,"Check List"] = "최상급 뽑기 11회 연출 및 카드 획득(고급 이상)"
            elif '뽑기' in y.pkgName:
                result.loc[i,"Check List"] = f"{y.pkgName} 연출 및 카드 획득"
            else:
                if len(y.itemList0) != 0 :
                    for item0 in y.itemList0 :
                        # if '다이아몬드' in str(item0):
                        #     result.loc[i,"Check List"] = "다이아몬드 이미지 노출"
                        #     i+=1
                        # else :
                        result.loc[i,"Check List"] = f"{item0} 인벤토리 획득"
                        i+=1
                else:
                    result.loc[i,"Check List"] = y.pkgName + " 인벤토리 획득"
            i += 1

    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
        for item0 in y.itemList0 :
            if '상자' in item0 :
                result.loc[i,"Category3"] = "상자 구성품\n획득 및 사용"
                break
            result.loc[i,"Category3"] = "아이템 사용"

        for item1 in y.itemList1 :
            if str(item1) != "nan" :
                item1 = item1.replace("코인[귀속]","코인")
                result.loc[i,"Check List"] = str(item1) + " 획득 및 사용 확인"
                i+=1
        
        i-=1
    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
        i += 1
        result.loc[i,"Category3"] = "구매 제한"
        
        if str(y.limit) == 'nan' or str(y.limit) == '무제한':
            result.loc[i,"Check List"] = "구매 제한 없음 (n회 구매 시, [구매 완료] 라벨 미노출)"
        else :            
            if '스텝' in y.pkgName :
                result.loc[i,"Check List"] = str(y.limit) + " 구매 시 스텝업 다음 단계 상품 노출\n(마지막 단계 : [구매 완료] 라벨 노출 및 터치 불가)"
            else :
                result.loc[i,"Check List"] = str(y.limit) + " 구매 시 [구매 완료] 라벨 노출 및 터치 불가"

        #i += 1
        #result.loc[i,"Check List"] = "상품 슬롯 하단에 [구매 완료] 라벨 노출"
    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
        curRow = i
        
        #print(f'{y.pkgName} {y.salesCheck}')
        if y.salesCheck == "판매 전" or y.salesCheck == "판매 시작":
            #print("추가")
            
            totalResult = pd.concat([totalResult,result], ignore_index=True)
            idList.append(y.pkgID)
        #print(len(totalResult))

        """상품별체크중단점"""

    totalResult = totalResult.replace("NaN","")
    totalResult = totalResult.replace("nan","")
    totalResult = totalResult.replace(np.nan,"")

    xlFileName = f"./{resultPath}/result_{time.strftime('%y%m%d_%H%M%S')}.xlsx"

    totalResult.to_excel(xlFileName, # directory and file name to write

                sheet_name = 'Sheet1', 

                na_rep = 'NaN', 

                float_format = "%.2f", 

                header = True, 

                #columns = ["group", "value_1", "value_2"], # if header is False

                index = True, 

                #index_label = "id", 

                startrow = 0, 

                startcol = 0, 

                #engine = 'xlsxwriter', 

                #freeze_panes = (2, 0)

                ) 

    return xlFileName


def write_data_cashshop_inspection(salesList : list[Sales], resultPath, check_box_list):

    '''
    유료상점_체크리스트_작성용\n
    check_box_list : 옵션 체크 리스트
    '''

    category_order = ['시즌 뽑기', '기간한정상품', '패키지', '카드', '재화', '이벤트']
    
    for sale in salesList:
        print(f'{sale.pkgName}|{sale.server}|{sale.salesCheck}|{sale.category}|{sale.order}')
    try:
        #    sale.salesCheck = float(sale.salesCheck)
        #salesList.sort(key=lambda a: (a.server, a.salesCheck, a.category, a.order))
        salesList.sort(key=lambda a: (
            a.salesCheck,
            a.server,
            category_order.index(a.category) if a.category in category_order else float('inf'),
            a.order
        ))

    except Exception as e:
        print(e)
        print("정렬에 문제 발생... 표에 제대로 입력됐는지 확인 필요...")

    # try:
    #     salesList.sort(key =lambda a: (a.server,a.salesCheck,a.category,a.order))
    # except Exception as e:
    #     print(e)
    #     print("정렬에 문제 발생... 표에 제대로 입력됐는지 확인 필요...")


    totalResult = pd.DataFrame()
    curRow = 0
    count = 0
    #tqdmCount0=0
    print("데이터 쓰는 중...")
    for y in tqdm(salesList):

        #tqdmCount0+=1
        y : Sales
        count += 1
        result = pd.DataFrame()


        if y.salesCheck == "판매 제외" or y.salesCheck == "판매 전"  :
            continue

        

        i = curRow
        result.loc[i,"Category1"] = y.server
        result.loc[i,"Category2"] = y.salesCheck
        result.loc[i,"Category3"] = y.pkgName

    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
        bonusStr = ""

        if int(y.bonus) == 0 :
            bonusStr =  "마일리지 X"
        else :            
            bonusStr =  f"{y.bonus} 마일리지"

        info_0 = f'{y.category}'
        info_1 = ""
        info_2 = ""
        info_2_1 = ""
        info_3 = ""
        
        if y.endDate == datetime.datetime.strptime("2099-12-31 00:00:00",'%Y-%m-%d %H:%M:%S') :
            info_expired = "상시 판매 상품(종료날짜 미표시)"
            if y.salesCheck != "판매 시작":
                continue
        else :
            info_expired = f" | {y.endDate.strftime('%m/%d/%Y(목) 점검 전 까지')}"

        if not check_box_list[1] or y.salesCheck == "판매 시작":
            info_0 = f'{info_0} / {y.price} / {bonusStr} / {y.limit}'

            info_expired = f"\n{info_expired}"

            if len(y.itemList0) != 0 :
                info_1 = "\n".join(y.itemList0)
                info_1 = info_1.replace("다이아몬드[귀속]","다이아몬드")
                info_1 = info_1.replace("다이아[귀속]","다이아")
                info_1 = info_1.replace("코인[귀속]","코인")
                info_1 = f'\n\n{info_1}'

                if len(y.itemList1) != 0 :
                    info_2 = "\n".join(map(str, y.itemList1))
                    info_2 = info_2.replace("nan\n","")
                    info_2 = info_2.replace("\n","\n- ")
                    if len(y.itemList0) != 0 :
                        info_2 = "\n\n사용 시 다음 아이템 획득\n- "+info_2
                    # elif len(y.itemList1) != 0 and len(y.itemList2) != 0 :
                    #     info_2 = "\n\n사용 시 다음 아이템 중 1종 획득\n- "+info_2
                    else:
                        info_2 = f'\n\n{info_2}'
                    
                    info_2 = info_2.replace("코인[귀속]","코인")

            else:
                info_2 = "\n".join(map(str, y.itemList1))
                info_2 = info_2.replace("nan\n","")
                #info_2 = info_2.replace("\n","\n- ")

                if len(y.itemList2) != 0 :
                    info_2_1 = "\n".join(map(str, y.itemList2))
                    info_2_1 = info_2_1.replace("nan\n","")
                    info_2_1 = info_2_1.replace("\n","\n- ")
                    #if len(y.itemList0) != 0 :
                    info_2_1 = "\n\n사용 시 다음 아이템 중 1종 획득\n- "+info_2_1

                    info_2 = f'\n\n{info_2}'
                    #else:
                        #info_2 = f'\n\n{info_2}'
                    
                    #info_2 = info_2.replace("코인[귀속]","코인")
            

            if check_box_list[3]:
                info_3 = f'\n\n* 상세정보 및 패키지 상자 구성품 내 [귀속] 노출 확인\n* 패키지 이미지 내 구성품 관련 이미지 노출 확인'

        result.loc[i,"Check List"] = f'{info_0}{info_expired}{info_1}{info_2}{info_2_1}{info_3}'

    # #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■
     
        result.loc[i,"pkgID"] = y.pkgID
        result.loc[i,"order"] = y.order


        #if y.salesCheck != "판매 제외" and y.salesCheck != "판매 전"  :
        totalResult = pd.concat([totalResult,result], ignore_index=True)

    totalResult = totalResult.replace("NaN","")
    totalResult = totalResult.replace("nan","")
    totalResult = totalResult.replace(np.nan,"")

    

    xlFileName = f"./{resultPath}/result_{time.strftime('%y%m%d_%H%M%S')}.xlsx"


    totalResult.to_excel(xlFileName, # directory and file name to write

                sheet_name = 'Sheet1', 

                na_rep = 'NaN', 

                float_format = "%.2f", 

                header = True, 

                #columns = ["group", "value_1", "value_2"], # if header is False

                index = True, 

                #index_label = "id", 

                startrow = 0, 

                startcol = 0, 

                #engine = 'xlsxwriter', 

                #freeze_panes = (2, 0)

                ) 
    
    return xlFileName


def postprocess_cashshop(xlFileName):
    wb = xl.load_workbook(xlFileName,data_only = True)
    sheetList = wb.sheetnames
    ws = wb[sheetList[0]]
    ws.column_dimensions['b'].width = 17
    ws.column_dimensions['c'].width = 17
    ws.column_dimensions['d'].width = 17
    ws.column_dimensions['e'].width = 50

    firstRow = 2
    lastRow = ws.max_row
    startRow_B =0
    startValue_B =""
    startRow_C = 0
    startRow_D = 0

    tqdmCount1 = 0
    print("엑셀 서식 처리중...")
    for i in tqdm(range(firstRow, lastRow+1)):
        tqdmCount1+=1
        #print(i)

        """서버 카테고리 합치기"""
        if (ws['b'+str(i)].value is not None) :
            if startRow_B == 0  :
                startRow_B = i
                startValue_B = ws['b'+str(i)].value
                #print(startRow_B)
            else :
                #firstTargetCell =  "C"+str(startRow_C)
                if ( ws['b'+str(i)].value != startValue_B) :
                    mergeTargetCell = "B"+str(startRow_B)+":B"+str(i-1)
                    ws.merge_cells(mergeTargetCell)
                    startValue_B = ws['b'+str(i)].value
                    startRow_B = i

        """판매 분류 카테고리 합치기"""
        if ws['c'+str(i)].value is not None:
            if startRow_C == 0 :
                startRow_C = i
                #print(startRow)
                startValue_C = ws['c'+str(i)].value
            else :
                # firstTargetCell =  "C"+str(startRow_C)
                # mergeTargetCell = "C"+str(startRow_C)+":C"+str(i-1)
                # ws.merge_cells(mergeTargetCell)
                # startRow_C = i
                if ( ws['c'+str(i)].value != startValue_C) :
                    mergeTargetCell = "c"+str(startRow_C)+":c"+str(i-1)
                    ws.merge_cells(mergeTargetCell)
                    startValue_C = ws['c'+str(i)].value
                    startRow_C = i


        if ws['d'+str(i)].value is not None:
            if startRow_D == 0 :
                startRow_D = i
                #print(startRow)
            else :
                firstTargetCell =  "D"+str(startRow_D)
                mergeTargetCell = "D"+str(startRow_D)+":D"+str(i-1)
                ws.merge_cells(mergeTargetCell)
                startRow_D = i


        ws['b'+str(i)].alignment = Alignment(
            horizontal='center'
            ,vertical='top'
            ,wrap_text=True)
        ws['b'+str(i)].font = Font(size = 9, bold = True)
        ws['c'+str(i)].alignment = Alignment(
            horizontal='center'
            ,vertical='top'
            ,wrap_text=True)
        ws['c'+str(i)].font = Font(size = 9, bold = True)
        ws['d'+str(i)].alignment = Alignment(
            horizontal='center'
            ,vertical='top'
            ,wrap_text=True)
        ws['d'+str(i)].font = Font(size = 9, bold = True)
        ws['e'+str(i)].alignment = Alignment(
            horizontal='left'
            ,vertical='top'
            ,wrap_text=True)
        ws['e'+str(i)].font = Font(size = 9, bold = False)
        
        #ws['e'+str(i)].value = process_temp_str(str(ws['e'+str(i)].value))


    #예외 마지막 셀병합
    ws.merge_cells("B"+str(startRow_B)+":B"+str(lastRow))
    ws.merge_cells("C"+str(startRow_C)+":C"+str(lastRow))
    ws.merge_cells("D"+str(startRow_D)+":D"+str(lastRow))

    #ws = highlight_belonging(ws)
    ws = find_and_replace(ws,"로얄 코인[귀속]","로얄 코인")
    ws = find_and_replace(ws,"로얄코인[귀속]","로얄 코인")
    ws = find_and_replace(ws,"다이아몬드[귀속]","다이아")
    ws = find_and_replace(ws,"[귀속][귀속]","[귀속]")
    ws = highlight_star_cells(ws)
    wb.save(xlFileName)


def process_temp_str(temp_str):
    # Define the font style for the red asterisk
    red_asterisk_font = Font(color=Color('FF0000'), bold=True)
    # Define the font style for the rest of the cell
    normal_font = Font(color=Color('000000'))

    # Split the temp_str value by newline characters
    temp_str_lines = temp_str.split("\n")

    # Process each line of the temp_str value
    result_lines = []
    for line in temp_str_lines:
        # Check if the line starts with an asterisk
        if line.startswith("*"):
            # Apply the red font to the asterisk and the rest of the line
            result_lines.append((line[:1] + line[1:].strip(), red_asterisk_font))
        else:
            # Apply the normal font to the entire line
            result_lines.append((line, normal_font))

    # Join the processed lines and return the result
    #return "\n".join(result_lines)
    #print(result_lines)
    return "\n".join([line[0] for line in result_lines])


def highlight_belonging(ws):
    # Define the font style for the highlighted text
    highlight_font = Font(color=Color('FF0000'), bold=True)

    # Iterate over all cells in the worksheet
    for row in ws.iter_rows():
        for cell in row:
            # Check if the cell contains the word "belonging"
            if "귀속" in str(cell.value).lower():
                # Split the cell value by the word "belonging"
                parts = str(cell.value).lower().split("귀속")
                # Concatenate the parts with the highlighted "belonging" in between
                highlighted_value = "귀속".join([part.strip() for part in parts])
                # Apply the highlight font to the "belonging" text
                cell.font = highlight_font
                # Set the cell value to the highlighted value
                cell.value = highlighted_value

    # Auto-fit the columns to adjust to the new cell values
    for column in ws.columns:
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].auto_size = True

    # Return the worksheet object
    return ws


def find_and_replace(ws, target_str, replace_str, font=None):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and target_str in str(cell.value):
                old_value = str(cell.value)
                new_value = old_value.replace(target_str, replace_str)
                cell.value = new_value
                
                if font:
                    font_obj = Font(name=font, bold=True, color="FF0000")
                    start_index = new_value.find(replace_str)
                    end_index = start_index + len(replace_str)
                    cell.font = Font(color="000000")
                    cell.font = font_obj
                    if start_index > 0:
                        cell.font = Font(color="000000", bold=False)
                    if end_index < len(new_value):
                        cell.font = Font(color="000000", bold=False)
    return ws


def highlight_star_cells(sheet):
    red_bold_font = Font(color="FF0000", bold=True)
    for row in sheet.iter_rows():
        for cell in row:
            try :
                if cell.value is not None and "*" in cell.value:
                    parts = cell.value.split("*")
                    new_parts = [f"{Font(color='black')}*{red_bold_font}".join(part.strip() for part in parts)]
                    cell.value = "\n".join(new_parts)
                    cell.font = red_bold_font
            except :
                continue
    return sheet


if __name__ == "__main__":
    #def make_process(self, result_path, data_file_name, contents_name, doctype, date_text, check_box_list):
    nation = 'KR'
    contents_name = "유료상점"
    doctype = "CheckList"
    doctype = "TestCase"
    
    result_path = f'{contents_name}_{doctype}'
    data_file_name = f'{contents_name}DATA_{nation} R2M.xlsx'#유료상점DATA_KR R2M.xlsx
    date_text = '2023-11-30'
    check_box_list = [True,True,False,False]
    
    if contents_name == "유료상점" : 
        if doctype == "CheckList" :
            data = extract_data_cashshop(data_file_name,date_text)
            result_file_name = write_data_cashshop_inspection(data,result_path,check_box_list)
            postprocess_cashshop(result_file_name)
        elif doctype == "TestCase" :
            data = extract_data_cashshop(data_file_name, date_text)
            result_file_name = write_data_cashshop(data,result_path)
            postprocess_cashshop(result_file_name)

    os.startfile(os.path.normpath(result_file_name))

#     #print("┃  R2M CASH SHOP CL MAKER  ┃")
#     print("체크리스트 타입 입력 :")
#     print("0:TC, 1:점검")
#     fileType = input(">:")
#     print("데이터파일명 입력 :")
#     print("0:국내, 1:대만")
#     countryType = input(">:")
#     #fileName = ""
#     if countryType == "0":
#         fileName = "유료상점DATA_KR.xlsx"
#     elif countryType == "1":
#         fileName = "유료상점DATA_TW.xlsx"
#     while not os.path.isfile(fileName) :
#         fileName = input(">:")
        

#     clType = ""
#     if fileType == "0":
#         clType = "TestCase"
#         print("업데이트날짜 입력 시, 해당 날짜 포함하여 이후 시작하는 상품만 작성")
#         print("YYYY-MM-DD")
#         print("(그냥 엔터키 입력 시, 오늘로 설정)")
#         tcStartDate = input(">: ")
#         if tcStartDate == "" :
#             tcStartDate = datetime.date.today().strftime('%Y-%m-%d')
#             #tcStartDate = "2000-01-01"
        
#     elif fileType == "1":
#         clType = "CheckList"
#         # 오늘 날짜 구하기
#         todayDate = datetime.datetime.today().date()

#         # 그 주의 점검 날짜 구하기 (대만:화, 국내:목)
#         dateID= 0
#         if countryType == "0":
#             dateID = (3,"목")
#         elif countryType == "1":
#             dateID = (1,"화")

#         days_until_target = (dateID[0] - todayDate.weekday()) % 7
#         thursdayDate = todayDate + datetime.timedelta(days=days_until_target)
#         tcStartDate = thursdayDate.strftime('%Y-%m-%d')

#         print(f"이번주 {dateID[1]}요일 {tcStartDate} 기준으로 작성됩니다.")



# #region basic Info

#     cashShopDir = f"./CashShop_{clType}"
#     if not os.path.isdir(cashShopDir) :
#         os.mkdir(cashShopDir)
#     tempDir = "./temp"
#     if not os.path.isdir(tempDir) :
#         os.mkdir(tempDir)


#     #xlFileName = f"./CL_CashShop_{clType}/result_{time.strftime('%y%m%d_%H%M%S')}.xlsx"
#     tempCsvName = f"./temp/tempCsv.csv"

# #endregion
#     try:
#         salesList = extract_data_cashshop(fileName,tcStartDate)
#         pass
#     except PermissionError:
#         print(f"해당 문서가 열려있습니다. 닫고 다시 시작해주세요. {fileName}")
#         input("아무 키나 누르세요...")
#         os.system('python ' + os.path.basename(__file__))


#     if fileType == "0":
#         xlFileName = write_data_cashshop(salesList)
#         postprocess_cashshop(xlFileName)
#     elif fileType == "1":
#         if salesList == None :
#             salesList = extract_data_cashshop(fileName,tcStartDate)
#         xlFileName = write_data_cashshop_inspection(salesList)
#         postprocess_cashshop(xlFileName)


#     print("생성완료")
#     for id in idList :
#         print(id)        
#     os.system('pause')


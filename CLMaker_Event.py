import pandas as pd
import time
import os
import gc
import re
import numpy as np
import openpyxl as xl
from openpyxl.styles import Font, Alignment, Color
from tqdm import tqdm
from openpyxl.styles.colors import BLACK
import datetime
from openpyxl.styles import Font, Color
#from openpyxl.styles.colors import RED
from openpyxl.utils import get_column_letter
from pathlib import Path
from enum import Enum, auto
#cashShopIdIndexList = cashShopIdList.index
clType = ""
cashShopDir = ""
tempDir = "./temp"
if not os.path.isdir(tempDir) :
    os.mkdir(tempDir)
tempCsvName = f"./temp/tempCsv.csv"

xlFileName = ""
tcStartDate = ""

class DocumentType(Enum) :
    TestCase = 0
    CheckList = 1
    

def dateCheck(start_date, end_date, today = datetime.date.today()):
    
    if start_date.date() == today :
        return "이벤트 시작"
    elif start_date.date() < today < end_date.date():
        return "이벤트 유지"
    elif today == end_date.date():
        return "이벤트 종료"
    # elif start_date.date() >= today - datetime.timedelta(days=1)  :
    #     return "판매 전"
    elif start_date.date() >= today  :
        return "이벤트 시작 전"
    else:
        return "이벤트 제외"

class Item():
    # def __init__(self) :
    #     self.id = ""
    #     self.name = ""
    #     self.count = ""

    def __init__(self, id, name, count, movelimit=6, removedate=0) :
        self.id = id
        self.name = name
        self.count = count
        self.movelimit = movelimit
        self.removedate = removedate
    # def additem(self, id, name, count) :
    #     self.id = id
    #     self.name = name
    #     self.count = count

class Craft():
    def __init__(self,price, limit, recipe, rate = 100) :
        self.price = price
        self.limit = limit
        self.recipe = recipe
        self.rate = rate

class Quest():
    def __init__(self, id, name) :
        self.id = id
        self.name = name
        
class Event():
    def __init__(self) :

        self.id = ""
        self.type = ""
        self.name = ""

        self.limit = ""
        self.server = ""
        self.start_date = ""
        self.end_date = ""

        self.desc_list = []
        self.item_list = []
        self.craft_list = []

        #self.item_list_1 = []#패스용
        #별도 저장값
        self.open_check = ""


def extract_data(fileName, tcStartDate):

    target = pd.DataFrame()
    

    sheet_names = pd.read_excel(fileName, sheet_name=None).keys()
    for sheet_name in sheet_names:
        
        if sheet_name == "사용법" or sheet_name == "템플릿" :
            continue
        
        curDf = pd.read_excel(fileName, sheet_name=sheet_name, na_values="")
        #target = target.append(curDf, ignore_index = True)
        target = pd.concat([target, curDf], ignore_index=True)
        del curDf
        gc.collect()

    target = target.replace('-',np.nan)
    targetIdList = target["ID"].dropna(axis=0) #실제 ID의 리스트
    targetIdIndexList = targetIdList.index #문서에서 ID가 입력된 행의 인덱스의 리스트

    totalCount = len(targetIdIndexList)
    targetList = []
    print("데이터 추출 중...")

    for j in tqdm(range(0,totalCount)):\
    

        if (j+1) >= len(targetIdIndexList) :
            tempDf = target[targetIdIndexList[j]:]
        else :
            tempDf = target[targetIdIndexList[j]:targetIdIndexList[j+1]]
        tempDf = tempDf.reset_index()
        

        a = Event()
        a.id = int(tempDf.loc[0,"ID"])
        a.type = str(tempDf.loc[0,"EventType"])
        a.name = str(tempDf.loc[0,"EventName"]) #+ "[귀속]"

        for k in range(len(tempDf)):
            str0 = tempDf.loc[k,'QuestID']
            str1 = tempDf.loc[k,'QuestDesc']
            if not pd.isnull(str1):
                if not pd.isnull(str0):
                    newQuest = Quest(str0,str1)
                    #a.desc_list.append(newQuest)
                else :
                    newQuest = Quest(0,str1)
                a.desc_list.append(newQuest)

            else :
                continue 
            

        for k in range(len(tempDf)):
            str0 = tempDf.loc[k,'ItemID_0']
            str1 = tempDf.loc[k,'ItemName_0']
            try:
                if '[귀속]' not in str1 :
                    str1 = f'{str1}[귀속]'
                    str1 = str1.replace(' [귀속]','[귀속]')
                if '코인' in str1 :
                    str1 = str1.replace('[귀속]','')
            except : 
                pass
            str2 = tempDf.loc[k,'ItemCount_0']
            str3 = tempDf.loc[k,'ItemMovelimit_0']
            try:
                str4 = pd.to_datetime(tempDf.loc[k,'ItemRemovedate_0'])
                str4 = str4.strftime('%Y-%m-%d')
            except:
                pass

            if not pd.isnull(str1):
                newItem = Item(str0,str1,str2,str3,str4)
                a.item_list.append(newItem)
            else :
                continue 


        for k in range(len(tempDf)):
            str0 = tempDf.loc[k,'CraftPrice']
            str1 = tempDf.loc[k,'CraftLimit']
            str2 = tempDf.loc[k,'CraftIngred']

            if not pd.isnull(str2):
                newCraft = Craft(str0,str1,str2)
                a.craft_list.append(newCraft)
                #a.item_list_0.append(f"{str1} {str2} | {str0}")
            else :
                continue 


        a.server = tempDf.loc[0,"Server"]
        a.limit = tempDf.loc[0,"Limit"]
       
        a.start_date = pd.to_datetime(tempDf.loc[0,"StartDate"])
        try: 
            a.end_date = pd.to_datetime(tempDf.loc[0,"EndDate"])
        except :
            str_end_date = str(tempDf.loc[0,"EndDate"])
            if str_end_date == "상시" :
                a.end_date = datetime.datetime.strptime("2099-12-31 00:00:00",'%Y-%m-%d %H:%M:%S')
            else :
                print(f"이벤트 종료 시간 입력 오류 : 'yyyy-mm-dd' 만 입력되어야 함. 시간이 써있지 않은지 확인 필요\n상품 : {a.id} | {a.name}" )
                os.system('pause')
                return
          

        # if "상시" in str(tempDf.loc[0,"EndDate"]) :
        #     a.endDate = datetime.datetime.strptime("2099-12-31 00:00:00",'%Y-%m-%d %H:%M:%S')
        # else: 
        #     a.endDate = pd.to_datetime(tempDf.loc[0,"EndDate"])
          
        #startDate = datetime.datetime.strptime(tcStartDate, '%Y-%m-%d')
        #a.salesCheck = dateCheck(a.startDate,a.endDate,startDate.date())
        startDate = datetime.datetime.strptime(tcStartDate, '%Y-%m-%d')
        a.open_check = dateCheck(a.start_date,a.end_date,startDate.date())


        #print(a)
        if targetList != None :
            targetList.append(a)
        else :
            targetList = a



        del a,tempDf
        gc.collect()

    return targetList


def write_data_event_testcase(targetList : list[Event], resultPath = "이벤트_TestCase"):
    totalResult = pd.DataFrame()
#print(len(salesList))

    targetList.sort(key =lambda a: (a.id))
    curRow = 0
    count = 0
    print("데이터 쓰는 중...")
    for y in tqdm(targetList):
        #y = Event()
        count += 1
        result = pd.DataFrame()

        i = curRow

        if y.open_check == "이벤트 유지" or y.open_check == "이벤트 종료"  :
            continue
    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■[출석]

        if y.type == "출석" :
            result.loc[i,"Category1"] = f'{y.type} 이벤트'
            result.loc[i,"Category2"] = f'{y.name}\n{y.id}'#y.pkgName + "\n" + str(y.pkgID)
            result.loc[i,"Category3"] = "이벤트명"
            result.loc[i,"Check List"] = f'{y.name}'
    #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            i += 1
            result.loc[i,"Category3"] = "기간"
            #result.loc[i,"Check List"] = f"{y.start_date.strftime('%Y-%m-%d')}({dateID[1]}) ~ {y.end_date.strftime('%Y-%m-%d')}({dateID[1]})"
            result.loc[i,"Check List"] = f"{y.start_date.strftime('%Y-%m-%d')} ~ {y.end_date.strftime('%Y-%m-%d')}"
    #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            i += 1            
            result.loc[i,"Category3"] = "리소스 이미지"
            result.loc[i,"Check List"] = '이벤트 리소스 적용'
            i += 1            
            result.loc[i,"Check List"] = '이벤트 리소스 내 기간 정상 노출'

    #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            i += 1
            result.loc[i,"Category3"] = "제한"
            result.loc[i,"Check List"] = f"{y.limit}"
    #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            for index, item in enumerate(y.item_list) :
                i += 1
                result.loc[i,"Category3"] = f'{index+1}일차'
                #result.loc[i,"Check List"] = f'{item.name}[귀속] {int(item.count)}개'.replace('[귀속][귀속]','[귀속]')
                
                if not pd.isnull(item.removedate) :
                    result.loc[i,"Check List"] = f'{item.name} {int(item.count)}개 (자동삭제 : {item.removedate}) 11:30'#.replace('[귀속][귀속]','[귀속]')
                else :
                    result.loc[i,"Check List"] = f'{item.name} {int(item.count)}개'#.replace('[귀속][귀속]','[귀속]')
                #result.loc[i,"Check List"] = f'{item.name} {int(item.count)}개'#.replace('[귀속][귀속]','[귀속]')


    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■[미션]
        elif y.type == "미션" :
            result.loc[i,"Category1"] = f'{y.type} 이벤트'
            result.loc[i,"Category2"] = f'{y.name}\n{y.id}'#y.pkgName + "\n" + str(y.pkgID)
            result.loc[i,"Category3"] = "이벤트명"
            result.loc[i,"Check List"] = f'{y.name}'
    #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            i += 1
            result.loc[i,"Category3"] = "기간"
            #result.loc[i,"Check List"] = f"{y.start_date.strftime('%Y-%m-%d')}({dateID[1]}) ~ {y.end_date.strftime('%Y-%m-%d')}({dateID[1]})"
            result.loc[i,"Check List"] = f"{y.start_date.strftime('%Y-%m-%d')} ~ {y.end_date.strftime('%Y-%m-%d')}"
        #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            i += 1            
            result.loc[i,"Category3"] = "리소스 이미지"
            result.loc[i,"Check List"] = '이벤트 리소스 적용'
            i += 1            
            result.loc[i,"Check List"] = '이벤트 리소스 내 기간 정상 노출'
#━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            i += 1
            result.loc[i,"Category3"] = "제한"
            result.loc[i,"Check List"] = f"{y.limit}"
    #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            for index, quest in enumerate(y.desc_list) :
                i += 1
                result.loc[i,"Category3"] = f'{quest.name}'
                #result.loc[i,"Check List"] = f'{item.name}[귀속] {int(item.count)}개'.replace('[귀속][귀속]','[귀속]')
                if not pd.isnull(item.removedate) :
                    result.loc[i,"Check List"] = f'{y.item_list[index].name} {int(y.item_list[index].count)}개 (자동삭제 : {item.removedate} 11:30)'#.replace('[귀속][귀속]','[귀속]')
                else:
                    result.loc[i,"Check List"] = f'{y.item_list[index].name} {int(y.item_list[index].count)}개'#.replace('[귀속][귀속]','[귀속]')

    #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            i += 1            
            result.loc[i,"Category3"] = "퀘스트 연결"
            result.loc[i,"Check List"] = '6개 퀘스트 완료 시, 메인 퀘스트 보상 획득 가능'

    #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    
    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■[드랍]
        elif y.type == "드랍" :
            result.loc[i,"Category1"] = f'{y.type} 이벤트'
            result.loc[i,"Category2"] = f'{y.name}\n{y.id}'#y.pkgName + "\n" + str(y.pkgID)
            result.loc[i,"Category3"] = "기간"
            #result.loc[i,"Check List"] = f"{y.start_date.strftime('%Y-%m-%d')}({dateID[1]}) ~ {y.end_date.strftime('%Y-%m-%d')}({dateID[1]})"
            result.loc[i,"Check List"] = f"{y.start_date.strftime('%Y-%m-%d')} 11:30:00 ~ {y.end_date.strftime('%Y-%m-%d')} 11:30:00(KST)"
    #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            i += 1            
            result.loc[i,"Category3"] = "드랍 위치"
            result.loc[i,"Check List"] = f'{y.desc_list[0].name}'
    #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            for index, item in enumerate(y.item_list) :
                i += 1
                result.loc[i,"Category3"] = "드랍 아이템"                
                result.loc[i,"Check List"] = f'{item.name} ({(item.id)})'.replace('.0','')
    #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            for index, item in enumerate(y.item_list) :
                i += 1
                result.loc[i,"Category3"] = "드랍 확률"                
                result.loc[i,"Check List"] = f'{item.name} : {round(item.count,2)}%'#.replace('[귀속][귀속]','[귀속]')

    
    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■[제작]
        elif y.type == "제작" :
            result.loc[i,"Category1"] = f'{y.type} 이벤트'


            for index, item in enumerate(y.item_list) :
                #i += 1
                try:
                    result.loc[i,"Category2"] = f'{item.name}\n{int(item.id)}'#y.pkgName + "\n" + str(y.pkgID)
                except:
                    result.loc[i,"Category2"] = f'{item.name}'#y.pkgName + "\n" + str(y.pkgID)

    #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                result.loc[i,"Category3"] = "제한"
                result.loc[i,"Check List"] = f"{y.craft_list[index].limit}"
    #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                i += 1            
                result.loc[i,"Category3"] = "기간"
                result.loc[i,"Check List"] = f"{y.start_date.strftime('%m/%d/%Y')} ~ {y.end_date.strftime('%m/%d/%Y')} (이벤트 기간 실데이터 2:30:00)"
    #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
       
                i += 1            
                result.loc[i,"Category3"] = "재료"
                result.loc[i,"Check List"] = f"{y.craft_list[index].recipe}"
    #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                i += 1            
                result.loc[i,"Category3"] = "비용"
                result.loc[i,"Check List"] = f"{int(y.craft_list[index].price)} 골드"
    #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                i += 1            
                result.loc[i,"Category3"] = "확률"
                result.loc[i,"Check List"] = f"{int(y.craft_list[index].rate)}%로 제작 성공"
    #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
                i += 1            
                result.loc[i,"Category3"] = "실제 제작"
                result.loc[i,"Check List"] = f"아이템 제작 버튼 터치 시, 인벤토리 내 '{item.name}' 획득"

                if not pd.isnull(item.removedate) :
                    i += 1
                    result.loc[i,"Check List"] = f"자동 삭제 '{item.removedate} 11:30' 적용"
                    i += 1    
    
    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■[도감]

        elif y.type == "도감" :
            result.loc[i,"Category1"] = f'{y.type} 이벤트'
            result.loc[i,"Category2"] = f'{y.name}\n{y.id}'#y.pkgName + "\n" + str(y.pkgID)
            result.loc[i,"Category3"] = "도감명"
            result.loc[i,"Check List"] = f'{y.name}'
    #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            i += 1
            result.loc[i,"Category3"] = "기간"
            #result.loc[i,"Check List"] = f"{y.start_date.strftime('%Y-%m-%d')}({dateID[1]}) ~ {y.end_date.strftime('%Y-%m-%d')}({dateID[1]})"
            result.loc[i,"Check List"] = f"이벤트 기간 : {y.start_date.strftime('%m/%d/%Y')} 11:30 ~ {y.end_date.strftime('%m/%d/%Y')} 11:30"
        #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            
            for index, quest in enumerate(y.desc_list) :
                i += 1
                result.loc[i,"Category3"] = f'필요 아이템'
                #result.loc[i,"Check List"] = f'{item.name}[귀속] {int(item.count)}개'.replace('[귀속][귀속]','[귀속]')
                result.loc[i,"Check List"] = f'{quest.name} ({int(quest.id)})'#.replace('[귀속][귀속]','[귀속]')

        #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            
            for index, item in enumerate(y.item_list) :
                i += 1
                result.loc[i,"Category3"] = f'능력치'
                try:
                    stat_amount = float(item.count)
                    if stat_amount < 1 :
                        stat_amount *= 100
                        stat_amount = f'{stat_amount}0%'
                    else :
                        stat_amount = int(stat_amount)

                    result.loc[i,"Check List"] = f"{item.name} +{stat_amount}".replace('[귀속]','')
                except : 
                    result.loc[i,"Check List"] = f'{item.name} +{(item.count)}'.replace('[귀속]','')#.replace('[귀속][귀속]','[귀속]')
    #■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■■[도감]

        elif y.type == "패스" :
            result.loc[i,"Category1"] = f'패스 이벤트'
            result.loc[i,"Category2"] = f'{y.name}\neventID : {y.id}'
            result.loc[i,"Category3"] = "패스 이름"
            result.loc[i,"Check List"] = f'{y.name}'
    #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            i += 1
            result.loc[i,"Category3"] = "기간"
            result.loc[i,"Check List"] = f"이벤트 기간 : {y.start_date.strftime('%m/%d/%Y')} 11:30 ~ {y.end_date.strftime('%m/%d/%Y')} 11:30"
        #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            
            for index, quest in enumerate(y.craft_list) :
                i += 1
                result.loc[i,"Category3"] = f'패스 미션 및 보상'
                result.loc[i,"Check List"] = f'{quest.limit} : {quest.recipe}'#.replace('[귀속][귀속]','[귀속]')

        #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
            
            for index, item in enumerate(y.item_list) :
                i += 1
                result.loc[i,"Category3"] = f'{y.desc_list[index].name}'
                result.loc[i,"Check List"] = f'{item.name} {int(item.count)}개'#.replace('[귀속][귀속]','[귀속]')
                # try:
                #     stat_amount = float(item.count)
                #     if stat_amount < 1 :
                #         stat_amount *= 100
                #         stat_amount = f'{stat_amount}0%'
                #     else :
                #         stat_amount = int(stat_amount)

                #     result.loc[i,"Check List"] = f"{item.name} +{stat_amount}".replace('[귀속]','')
                # except : 
                #     result.loc[i,"Check List"] = f'{item.name} +{(item.count)}'.replace('[귀속]','')#.replace('[귀속][귀속]','[귀속]')

    #     if len(y.itemList0) != 0 :
    #         desc0 = "\n".join(y.itemList0)
    #         desc0 = desc0.replace("다이아몬드[귀속]","다이아몬드")
    #         result.loc[i,"Check List"] = desc0
    #     else : 
    #         result.loc[i,"Check List"] = f'{y.pkgName} 상자[귀속]'

    #     i += 1
    #     desc1 = "\n".join(map(str, y.itemList1))
    #     desc1 = desc1.replace("nan\n","")
    #     desc1 = desc1.replace("\n","\n- ")
    #     result.loc[i,"Check List"] = "사용 시 다음 아이템 획득\n\n- "+desc1

    #     i += 1
    #     desc0 = desc0.replace("\n","\n- ")
    #     result.loc[i,"Check List"] = "<"+y.pkgName+"> 구성품 상세 정보\n- " + desc0

    # #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    #     i += 1
    #     result.loc[i,"Category3"] = "상품슬롯"

    #     for item1 in y.itemList1 :
    #         if str(item1) != "nan" :
    #             result.loc[i,"Check List"] = str(item1) + " 관련 이미지 노출"
    #             i+=1

    #     if int(y.bonus) == 0 :
    #         result.loc[i,"Check List"] =  "마일리지 미노출"
    #     else :            
    #         result.loc[i,"Check List"] =  "마일리지 : " + str(y.bonus)+ " 적립"
    #     i+=1
    #     result.loc[i,"Check List"] = "구매 제한 : " + str(y.limit)
    #     i+=1
    #     result.loc[i,"Check List"] = "구매 가격 : " + y.price
        
    # #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    #     i += 1
    #     result.loc[i,"Category3"] = "아이템 구매"

    #     if "원" in y.price or "TWD" in y.price:
    #         result.loc[i,"Check List"] = f"결제 모듈 내 {y.pkgName} 노출"
    #         i += 1
    #         result.loc[i,"Check List"] = f"결제 모듈 내 {y.price} 노출"
    #         i += 1
    #         result.loc[i,"Check List"] = f"결제 완료 시 보관함으로 획득"
    #     else :
    #         result.loc[i,"Check List"] = y.price + " 차감"
    # #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

    #     i += 1
    #     result.loc[i,"Category3"] = "마일리지"

    #     if int(y.bonus) == 0 :
    #         result.loc[i,"Check List"] =  "미노출"
    #     else :            
    #         result.loc[i,"Check List"] =  "마일리지 : " + str(y.bonus)+ " 적립"

    # #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    #     i += 1
    #     result.loc[i,"Category3"] = "아이템 획득"
    #     result.loc[i,"Check List"] = y.pkgName + " 상자[귀속] 인벤토리 획득"
    # #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    #     i += 1
    #     result.loc[i,"Category3"] = "아이템 사용"

    #     for item1 in y.itemList1 :
    #         if str(item1) != "nan" :
    #             result.loc[i,"Check List"] = "- " + str(item1) + " 획득 및 사용"
    #             i+=1
        
    #     i-=1
    # #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
    #     i += 1
    #     result.loc[i,"Category3"] = "구매 제한"
    #     result.loc[i,"Check List"] = str(y.limit) + " 구매 시 상품 슬롯 비활성화"
    #     i += 1
    #     result.loc[i,"Check List"] = "상품 슬롯 하단에 [구매 완료] 라벨 노출"
    # #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        curRow = i
        
        #print(f'{y.pkgName} {y.salesCheck}')
        if y.open_check == "이벤트 시작 전" or y.open_check == "이벤트 시작":
            #print("추가")
            
            totalResult = pd.concat([totalResult,result], ignore_index=True)
        #print(len(totalResult))

    totalResult = totalResult.replace("NaN","")
    totalResult = totalResult.replace("nan","")
    totalResult = totalResult.replace(np.nan,"")

    xlFileName = f"./{resultPath}/result_{time.strftime('%y%m%d_%H%M%S')}.xlsx"

    totalResult.to_excel(xlFileName, # directory and file name to write

                sheet_name = 'Sheet1', 

                na_rep = 'NaN', 

                float_format = "%.2f", 

                header = True, 

                #columns = ["group", "value_1", "value_2"], # if header is False

                index = True, 

                #index_label = "id", 

                startrow = 0, 

                startcol = 0, 

                #engine = 'xlsxwriter', 

                #freeze_panes = (2, 0)

                )
    
    return xlFileName


def write_data(targetList : list[Event], resultPath = "이벤트_CheckList", options = None):

    targetList.sort(key =lambda a: (a.open_check, a.id))

    totalResult = pd.DataFrame()
    curRow = 0
    count = 0
    print("데이터 쓰는 중...")
    for y in tqdm(targetList):

        count += 1
        result = pd.DataFrame()


        if y.open_check == "이벤트 제외" or y.open_check == "이벤트 전"  :
            continue

        

        i = curRow
        result.loc[i,"Category1"] = y.server
        result.loc[i,"Category2"] = y.open_check
        result.loc[i,"Category3"] = f'{y.name}\n[{y.type}]'

    #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        #if y.open_check == "이벤트 시작" :

        # dateID= 0
        # if data_filename_input == "0":
        #     dateID = (3,"목")
        # elif data_filename_input == "1":
       # dateID = (1,"화")
       
        if options[1] : #CL 유지/종료 항목 내용 생략
            if y.open_check == "이벤트 유지" :
                if not options[2] and y.end_date == datetime.datetime.strptime("2099-12-31 00:00:00",'%Y-%m-%d %H:%M:%S'):
                    continue
    
        info = ""
        if y.type != "도감":
            #info = f"{y.start_date.strftime('%Y-%m-%d')}({dateID[1]}) ~ {y.end_date.strftime('%Y-%m-%d')}({dateID[1]})\n"
            info = f"{y.start_date.strftime('%Y-%m-%d')} ~ {y.end_date.strftime('%Y-%m-%d')}\n"
        
        if y.type == "출석" or y.type == "미션" :
            info += f"{y.limit} 수행 가능\n"
        
        if y.type == "출석" :
            for index, item in enumerate(y.item_list) :
                if not pd.isnull(item.removedate) :
                    info += f'\n{int(index)+1}일차 : {item.name} {int(item.count)}개 (자동삭제 : {item.removedate} 11:30)'
                else :
                    info += f'\n{int(index)+1}일차 : {item.name} {int(item.count)}개'
        elif y.type == "미션" :
            for index, item in enumerate(y.item_list) :
                if not pd.isnull(item.removedate) :
                    info += f'\n{y.desc_list[index].name} : {item.name} {int(item.count)}개 (자동삭제 : {item.removedate} 11:30)'
                else :
                    info += f'\n{y.desc_list[index].name} : {item.name} {int(item.count)}개'
        elif y.type == "드랍" :
            for index, item in enumerate(y.item_list) :
                info += f'\n{y.desc_list[index].name} : {item.name} {round(item.count,2)}%'
        elif y.type == "제작" :
            for index, item in enumerate(y.item_list) :
                craft = y.craft_list[index]
                info += f'\n{item.name} ({craft.limit})\n골드 {int(craft.price)} + {craft.recipe}\n'
        elif y.type == "도감" :
            info += f"이벤트 기간 : {y.start_date.strftime('%m/%d/%Y')} 11:30 ~ {y.end_date.strftime('%m/%d/%Y')} 11:30\n"
            for index, desc in enumerate(y.desc_list) :
                info += f'\n재료 : {desc.name}\n'
            for index, item in enumerate(y.item_list) :
                if item.count >= 1 :
                    item.count = int(item.count)
                else : 
                    item.count = f'{float(item.count) * 100}%'
                info += f'\n{item.name} +{(item.count)}'
                info = info.replace('[귀속]','')
                
        if options[1] : #CL 유지/종료 항목 내용 생략
            if y.open_check == "이벤트 유지" :
                info = f'{y.type} {y.open_check} 확인 | 종료일 : {y.end_date.strftime("%m/%d/%Y 11:30")}'
            elif y.open_check == "이벤트 종료" : 
                info = f'{y.type} {y.open_check} 확인'


        result.loc[i,"Check List"] = f'{info}'

    # #━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
     
        result.loc[i,"ETC"] = y.id

        del y
        gc.collect()
        #if y.salesCheck != "판매 제외" and y.salesCheck != "판매 전"  :
        totalResult = pd.concat([totalResult,result], ignore_index=True)

    totalResult = totalResult.replace("NaN","")
    totalResult = totalResult.replace("nan","")
    totalResult = totalResult.replace(np.nan,"")

    xlFileName = f"./{resultPath}/result_{time.strftime('%y%m%d_%H%M%S')}.xlsx"

    totalResult.to_excel(xlFileName, # directory and file name to write

                sheet_name = 'Sheet1', 

                na_rep = 'NaN', 

                float_format = "%.2f", 

                header = True, 

                #columns = ["group", "value_1", "value_2"], # if header is False

                index = True, 

                #index_label = "id", 

                startrow = 0, 

                startcol = 0, 

                #engine = 'xlsxwriter', 

                #freeze_panes = (2, 0)

                ) 

    return xlFileName

def postprocess_cashshop(xlFileName):
    wb = xl.load_workbook(xlFileName,data_only = True)
    sheetList = wb.sheetnames
    ws = wb[sheetList[0]]
    ws.column_dimensions['b'].width = 17
    ws.column_dimensions['c'].width = 17
    ws.column_dimensions['d'].width = 17
    ws.column_dimensions['e'].width = 50

    firstRow = 2
    lastRow = ws.max_row
    startRow_B =0
    startValue_B =""
    startRow_C = 0
    startRow_D = 0

    tqdmCount1 = 0
    print("엑셀 서식 처리중...")
    for i in tqdm(range(firstRow, lastRow+1)):
        tqdmCount1+=1
        #print(i)
        if (ws['b'+str(i)].value is not None) :
            if startRow_B == 0  :
                startRow_B = i
                startValue_B = ws['b'+str(i)].value
                #print(startRow_B)
            else :
                #firstTargetCell =  "C"+str(startRow_C)
                if ( ws['b'+str(i)].value != startValue_B) :
                    mergeTargetCell = "B"+str(startRow_B)+":B"+str(i-1)
                    ws.merge_cells(mergeTargetCell)
                    startValue_B = ws['b'+str(i)].value
                    startRow_B = i

        if ws['c'+str(i)].value is not None:
            if startRow_C == 0 :
                startRow_C = i
                #print(startRow)
                startValue_C = ws['c'+str(i)].value
            else :
                # firstTargetCell =  "C"+str(startRow_C)
                # mergeTargetCell = "C"+str(startRow_C)+":C"+str(i-1)
                # ws.merge_cells(mergeTargetCell)
                # startRow_C = i
                if ( ws['C'+str(i)].value != startValue_C) :
                    mergeTargetCell = "c"+str(startRow_C)+":c"+str(i-1)
                    ws.merge_cells(mergeTargetCell)
                    startValue_C = ws['C'+str(i)].value
                    startRow_C = i


        if ws['d'+str(i)].value is not None:
            if startRow_D == 0 :
                startRow_D = i
                #print(startRow)
                '''추가 230818'''
                startValue_D = ws['d'+str(i)].value
            else :
                '''원래'''
                # firstTargetCell =  "D"+str(startRow_D)
                # mergeTargetCell = "D"+str(startRow_D)+":D"+str(i-1)
                # ws.merge_cells(mergeTargetCell)
                # startRow_D = i
                '''변경 230818'''

                if ( ws['d'+str(i)].value != startValue_D) :
                    mergeTargetCell = "D"+str(startRow_D)+":D"+str(i-1)
                    ws.merge_cells(mergeTargetCell)
                    startValue_D = ws['D'+str(i)].value
                    startRow_D = i

        ws['b'+str(i)].alignment = Alignment(
            horizontal='center'
            ,vertical='top'
            ,wrap_text=True)
        ws['b'+str(i)].font = Font(size = 9, bold = True)
        ws['c'+str(i)].alignment = Alignment(
            horizontal='center'
            ,vertical='top'
            ,wrap_text=True)
        ws['c'+str(i)].font = Font(size = 9, bold = True)
        ws['d'+str(i)].alignment = Alignment(
            horizontal='center'
            ,vertical='top'
            ,wrap_text=True)
        ws['d'+str(i)].font = Font(size = 9, bold = True)
        ws['e'+str(i)].alignment = Alignment(
            horizontal='left'
            ,vertical='top'
            ,wrap_text=True)
        ws['e'+str(i)].font = Font(size = 9, bold = False)
        
        #ws['e'+str(i)].value = process_temp_str(str(ws['e'+str(i)].value))


    #예외 마지막 셀병합
    ws.merge_cells("B"+str(startRow_B)+":B"+str(lastRow))
    ws.merge_cells("C"+str(startRow_C)+":C"+str(lastRow))
    ws.merge_cells("D"+str(startRow_D)+":D"+str(lastRow))

    #ws = highlight_belonging(ws)
    #ws = find_and_replace(ws,"귀속","귀속")
    ws = highlight_star_cells(ws)
    wb.save(xlFileName)


def process_temp_str(temp_str):
    # Define the font style for the red asterisk
    red_asterisk_font = Font(color=Color('FF0000'), bold=True)
    # Define the font style for the rest of the cell
    normal_font = Font(color=Color('000000'))

    # Split the temp_str value by newline characters
    temp_str_lines = temp_str.split("\n")

    # Process each line of the temp_str value
    result_lines = []
    for line in temp_str_lines:
        # Check if the line starts with an asterisk
        if line.startswith("*"):
            # Apply the red font to the asterisk and the rest of the line
            result_lines.append((line[:1] + line[1:].strip(), red_asterisk_font))
        else:
            # Apply the normal font to the entire line
            result_lines.append((line, normal_font))

    # Join the processed lines and return the result
    #return "\n".join(result_lines)
    #print(result_lines)
    return "\n".join([line[0] for line in result_lines])


def highlight_belonging(ws):
    # Define the font style for the highlighted text
    highlight_font = Font(color=Color('FF0000'), bold=True)

    # Iterate over all cells in the worksheet
    for row in ws.iter_rows():
        for cell in row:
            # Check if the cell contains the word "belonging"
            if "귀속" in str(cell.value).lower():
                # Split the cell value by the word "belonging"
                parts = str(cell.value).lower().split("귀속")
                # Concatenate the parts with the highlighted "belonging" in between
                highlighted_value = "귀속".join([part.strip() for part in parts])
                # Apply the highlight font to the "belonging" text
                cell.font = highlight_font
                # Set the cell value to the highlighted value
                cell.value = highlighted_value

    # Auto-fit the columns to adjust to the new cell values
    for column in ws.columns:
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].auto_size = True

    # Return the worksheet object
    return ws


def find_and_replace(ws, target_str, replace_str, font=None):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and target_str in str(cell.value):
                old_value = str(cell.value)
                new_value = old_value.replace(target_str, replace_str)
                cell.value = new_value
                
                if font:
                    font_obj = Font(name=font, bold=True, color="FF0000")
                    start_index = new_value.find(replace_str)
                    end_index = start_index + len(replace_str)
                    cell.font = Font(color="000000")
                    cell.font = font_obj
                    if start_index > 0:
                        cell.font = Font(color="000000", bold=False)
                    if end_index < len(new_value):
                        cell.font = Font(color="000000", bold=False)
    return ws


def highlight_star_cells(sheet):
    red_bold_font = Font(color="FF0000", bold=True)
    for row in sheet.iter_rows():
        for cell in row:
            try :
                if cell.value is not None and "*" in cell.value:
                    parts = cell.value.split("*")
                    new_parts = [f"{Font(color='black')}*{red_bold_font}".join(part.strip() for part in parts)]
                    cell.value = "\n".join(new_parts)
                    cell.font = red_bold_font
            except :
                continue
    return sheet


if __name__ == "__main__":

    # while True:
    #     file_type_input = input("문서타입(0:TC, 1:CL) >: ")
    #     try : 
    #         file_type = DocumentType(int(file_type_input)).name
    #         break
    #     except :             
    #         print("잘못된 입력입니다. 다시 입력해주세요.")

    # file_dict = {"0": "이벤트DATA_KR.xlsx", "1": "이벤트DATA_TW.xlsx"}
    # # 파일명 입력 받기
    # while True:
    #     data_filename_input = input("국가설정(0:KR, 1:TW): ")
    #     if data_filename_input in file_dict:
    #         data_filename = file_dict[data_filename_input]
    #         if Path(data_filename).is_file():
    #             break
    #         else:
    #             print("파일이 존재하지 않습니다. 다시 입력해주세요.")
    #     else:
    #         print("잘못된 입력입니다. 다시 입력해주세요.")

    # todayDate = datetime.datetime.today().date()

    # # 그 주의 점검 날짜 구하기 (대만:화, 국내:목)
    # if file_type == DocumentType.TestCase :

    #     check_start_date = input("업데이트날짜(YYYY-MM-DD) >: ")
    #     if check_start_date == "" :
    #         check_start_date = datetime.datetime.now().strftime('%Y-%m-%d')

    # else:
    #     global dateID
    #     dateID= 0
    #     if data_filename_input == "0":
    #         dateID = (3,"목")
    #     elif data_filename_input == "1":
    #         dateID = (1,"화")

    #     days_until_target = (dateID[0] - todayDate.weekday()) % 7
    #     thursdayDate = todayDate + datetime.timedelta(days=days_until_target)
    #     check_start_date = thursdayDate.strftime('%Y-%m-%d')

    #     print(f"이번주 {dateID[1]}요일 {check_start_date} 기준으로 작성됩니다.")

    # #check_start_date = datetime.datetime.strptime(check_start_date, '%Y-%m-%d')

    # result_directory = f"./Event_{file_type}"
    # if not os.path.isdir(result_directory) :
    #     os.mkdir(result_directory)

    # #xl_filename = f"{result_directory}/result_{time.strftime('%y%m%d_%H%M%S')}.xlsx"

    # targetList = extract_data(data_filename,check_start_date)
    # xlFileName = write_data(targetList)
    # postprocess_cashshop(xlFileName)


    # print("complete!")
    # os.system("pause")
    nation = 'KR'
    doctype = "TestCase"

    contents_name = "이벤트"
    
    result_path = f'{contents_name}_{doctype}'
    data_file_name = f'{contents_name}DATA_{nation} R2M.xlsx'#유료상점DATA_KR R2M.xlsx
    date_text = '2023-12-07'
    check_box_list = [True,True,False,True]

    if doctype == "CheckList" :
        data = extract_data(data_file_name, date_text)
        result_file_name = write_data(data,result_path,check_box_list)
    elif doctype == "TestCase" :
        data = extract_data(data_file_name,date_text)
        result_file_name = write_data_event_testcase(data,result_path)


    postprocess_cashshop(result_file_name)

    os.startfile(os.path.normpath(result_file_name))